"""
Run unit tests for all modules and write an Excel report.

  python tests/run_tests.py              # run all test suites
  python tests/run_tests.py config       # config tests only
  python tests/run_tests.py evaluation   # evaluation tests only

Reports are written to tests/completed/Unit_Tests_<timestamp>.xlsx
"""
from __future__ import annotations

import os
import sys
import time
from datetime import datetime
from pathlib import Path

# Bootstrap path so all packages are importable
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pytest
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

TESTS_DIR = PROJECT_ROOT / "tests"
COMPLETED = TESTS_DIR / "completed"

# Each suite maps to a test subdirectory
TEST_SUITES: dict[str, dict] = {
    "config":        {"label": "Config",             "folder": str(TESTS_DIR / "test_config"),        "color": "1565C0"},
    "evaluation":    {"label": "Evaluation Metrics",  "folder": str(TESTS_DIR / "test_evaluation"),    "color": "00695C"},
    "experiments":   {"label": "Experiment Metrics",  "folder": str(TESTS_DIR / "test_experiments"),   "color": "4A148C"},
    "knowledge_base":{"label": "Knowledge Base",      "folder": str(TESTS_DIR / "test_knowledge_base"),"color": "BF360C"},
    "retrieval":     {"label": "Retrieval Pipelines", "folder": str(TESTS_DIR / "test_retrieval"),     "color": "1B5E20"},
    "golden_dataset":{"label": "Golden Dataset",      "folder": str(TESTS_DIR / "test_golden_dataset"),"color": "E65100"},
}


# -- pytest result collector ---------------------------------------------------

class _Collector:
    """Minimal pytest plugin that records per-test outcomes."""

    def __init__(self) -> None:
        self.results: list[dict] = []

    def pytest_runtest_logreport(self, report: pytest.TestReport) -> None:
        if report.when == "call":
            self.results.append({
                "nodeid":   report.nodeid,
                "name":     report.nodeid.split("::")[-1],
                "file":     report.nodeid.split("::")[0],
                "status":   "PASSED" if report.passed else ("FAILED" if report.failed else "SKIPPED"),
                "duration": round(getattr(report, "duration", 0.0), 4),
                "error":    _fmt_error(report.longrepr) if report.failed else "",
            })
        elif report.when == "setup" and report.failed:
            self.results.append({
                "nodeid":   report.nodeid,
                "name":     report.nodeid.split("::")[-1],
                "file":     report.nodeid.split("::")[0],
                "status":   "ERROR",
                "duration": round(getattr(report, "duration", 0.0), 4),
                "error":    _fmt_error(report.longrepr),
            })


def _fmt_error(longrepr) -> str:
    if longrepr is None:
        return ""
    try:
        text = str(longrepr)
        lines = [ln for ln in text.splitlines() if ln.strip()]
        return " | ".join(lines[-3:])[:500]
    except Exception:
        return str(longrepr)[:500]


# -- Excel writer --------------------------------------------------------------

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_FILL = {
    "PASSED":  PatternFill("solid", fgColor="C8E6C9"),
    "FAILED":  PatternFill("solid", fgColor="FFCDD2"),
    "ERROR":   PatternFill("solid", fgColor="FFE0B2"),
    "SKIPPED": PatternFill("solid", fgColor="E3F2FD"),
}


def _lighten(hex_color: str) -> str:
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * 0.88)
    g = int(g + (255 - g) * 0.88)
    b = int(b + (255 - b) * 0.88)
    return f"{r:02X}{g:02X}{b:02X}"


def _write_summary_sheet(
    wb: Workbook,
    suite_results: dict[str, list[dict]],
    suite_meta: dict[str, dict],
    total_elapsed: float,
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "RAG Energy Forecasting - Unit Test Report"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="212121")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Timestamp
    ws.merge_cells("A2:G2")
    ts_cell = ws["A2"]
    ts_cell.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}  |  "
        f"Total run time: {total_elapsed:.1f}s"
    )
    ts_cell.font = Font(italic=True, size=10, color="555555")
    ts_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = ["Module", "Total", "Passed", "Failed", "Errors", "Skipped", "Pass Rate"]
    col_widths = [22, 10, 10, 10, 10, 10, 12]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="37474F")

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    grand_total = grand_pass = grand_fail = grand_err = grand_skip = 0

    for row_offset, (key, results) in enumerate(suite_results.items(), 4):
        label = suite_meta[key]["label"]
        color = suite_meta[key]["color"]
        total = len(results)
        passed = sum(1 for r in results if r["status"] == "PASSED")
        failed = sum(1 for r in results if r["status"] == "FAILED")
        errors = sum(1 for r in results if r["status"] == "ERROR")
        skipped = sum(1 for r in results if r["status"] == "SKIPPED")
        rate = f"{passed / total * 100:.1f}%" if total else "N/A"

        grand_total += total
        grand_pass += passed
        grand_fail += failed
        grand_err += errors
        grand_skip += skipped

        row_fill = PatternFill("solid", fgColor=_lighten(color))
        values = [label, total, passed, failed, errors, skipped, rate]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 1:
                cell.font = Font(bold=True, color=color, size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Totals row
    total_row = 4 + len(suite_results)
    grand_rate = f"{grand_pass / grand_total * 100:.1f}%" if grand_total else "N/A"
    totals = ["TOTAL", grand_total, grand_pass, grand_fail, grand_err, grand_skip, grand_rate]
    tfill = PatternFill("solid", fgColor="ECEFF1")
    for col_idx, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.fill = tfill
        cell.font = Font(bold=True, size=11)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A4"


def _write_detail_sheet(
    wb: Workbook, label: str, results: list[dict], hex_color: str
) -> None:
    ws = wb.create_sheet(title=label[:31])  # Excel sheet name limit

    headers = ["#", "Test Name", "File", "Status", "Duration (s)", "Error"]
    col_widths = [5, 50, 40, 10, 14, 70]

    hfill = PatternFill("solid", fgColor=hex_color)
    hfont = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, r in enumerate(results, 2):
        values = [row_idx - 1, r["name"], r["file"], r["status"], r["duration"], r["error"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 3, 6)))
            if col_idx == 4:
                cell.fill = _STATUS_FILL.get(r["status"], PatternFill())
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


# -- runner --------------------------------------------------------------------

def run_suite(key: str, meta: dict) -> tuple[list[dict], int]:
    """Run pytest on one test folder and return (results, exit_code)."""
    folder = meta["folder"]
    if not Path(folder).exists():
        print(f"  [SKIP] {folder} does not exist")
        return [], 5

    collector = _Collector()
    exit_code = pytest.main(
        [folder, "-v", "--tb=short", "--no-header", "-q"],
        plugins=[collector],
    )
    return collector.results, int(exit_code)


def main(selected: list[str]) -> None:
    suites_to_run = {k: v for k, v in TEST_SUITES.items() if k in selected}
    if not suites_to_run:
        print(f"Unknown suite(s): {selected}. Valid: {list(TEST_SUITES)}")
        sys.exit(1)

    suite_results: dict[str, list[dict]] = {}
    any_failed = False
    wall_start = time.time()

    for key, meta in suites_to_run.items():
        print(f"\n{'=' * 60}")
        print(f"  Running {meta['label']} tests")
        print(f"{'=' * 60}")
        t0 = time.time()
        results, exit_code = run_suite(key, meta)
        elapsed = time.time() - t0

        passed = sum(1 for r in results if r["status"] == "PASSED")
        print(f"\n  {meta['label']}: {passed}/{len(results)} passed  ({elapsed:.1f}s)")
        if exit_code not in (0, 5):
            any_failed = True
        suite_results[key] = results

    total_elapsed = time.time() - wall_start

    # Write Excel report
    COMPLETED.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, suite_results, TEST_SUITES, total_elapsed)
    for key, results in suite_results.items():
        if results:
            _write_detail_sheet(wb, TEST_SUITES[key]["label"], results, TEST_SUITES[key]["color"])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = COMPLETED / f"Unit_Tests_{timestamp}.xlsx"
    wb.save(report_path)

    print(f"\n{'=' * 60}")
    print(f"  Excel report saved -> {report_path.relative_to(PROJECT_ROOT)}")
    print(f"  Total time: {total_elapsed:.1f}s")
    print(f"{'=' * 60}\n")

    sys.exit(1 if any_failed else 0)


if __name__ == "__main__":
    args = sys.argv[1:]
    selected = args if args else list(TEST_SUITES.keys())
    main(selected)
